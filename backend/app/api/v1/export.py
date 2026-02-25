from uuid import UUID
import zipfile
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import csv
import io
from app.database import get_db
from app.dependencies import get_current_user
from app.models.user import User
from app.models.job import ScrapingJob, ScrapedProfile, ScrapedPost, PageAuthorProfile
from app.schemas.export import ExportRequest, FacebookAdsExportRequest

router = APIRouter()


async def _get_author(db: AsyncSession, job_id: UUID) -> PageAuthorProfile | None:
    result = await db.execute(
        select(PageAuthorProfile).where(PageAuthorProfile.job_id == job_id)
    )
    return result.scalar_one_or_none()


# Standard 20-field format
FIELDNAMES = [
    "ID", "Name", "First Name", "Last Name",
    "Gender", "Birthday", "Phone", "Relationship", "Education", "Work",
    "Position", "Hometown", "Location", "Website", "Languages",
    "UsernameLink", "Username", "About", "Picture URL", "Updated Time",
]

# Facebook Ads Manager custom audience format
FB_ADS_FIELDNAMES = [
    "email", "phone", "fn", "ln", "ct", "st", "country", "dob", "gender", "zip",
]


@router.get("/{job_id}/csv")
async def export_csv(
    job_id: UUID,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Verify job ownership
    job_result = await db.execute(
        select(ScrapingJob).where(
            ScrapingJob.id == job_id,
            ScrapingJob.tenant_id == user.tenant_id,
        )
    )
    job = job_result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Get profiles
    result = await db.execute(
        select(ScrapedProfile)
        .where(ScrapedProfile.job_id == job_id, ScrapedProfile.scrape_status == "success")
    )
    profiles = result.scalars().all()

    # Generate CSV
    output = io.StringIO()

    # Author info header
    author = await _get_author(db, job_id)
    if author:
        output.write(f"# Page: {author.name or ''} ({author.platform_object_id})")
        if author.category:
            output.write(f" | Category: {author.category}")
        if author.location:
            output.write(f" | Location: {author.location}")
        output.write("\n")

    writer = csv.DictWriter(output, fieldnames=FIELDNAMES)
    writer.writeheader()

    for p in profiles:
        writer.writerow({
            "ID": p.platform_user_id,
            "Name": p.name or "NA",
            "First Name": p.first_name or "NA",
            "Last Name": p.last_name or "NA",
            "Gender": p.gender or "NA",
            "Birthday": p.birthday or "NA",
            "Phone": p.phone or "NA",
            "Relationship": p.relationship_status or "NA",
            "Education": p.education or "NA",
            "Work": p.work or "NA",
            "Position": p.position or "NA",
            "Hometown": p.hometown or "NA",
            "Location": p.location or "NA",
            "Website": p.website or "NA",
            "Languages": p.languages or "NA",
            "UsernameLink": p.username_link or "NA",
            "Username": p.username or "NA",
            "About": p.about or "NA",
            "Picture URL": p.picture_url or "NA",
            "Updated Time": p.scraped_at.strftime("%Y-%m-%d %H:%M:%S") if p.scraped_at else "NA",
        })

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=socybase_export_{job_id}.csv"},
    )


@router.get("/{job_id}/facebook-ads")
async def export_facebook_ads(
    job_id: UUID,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export in same format as CSV/XLSX (full 20-field profile data)."""
    job_result = await db.execute(
        select(ScrapingJob).where(
            ScrapingJob.id == job_id,
            ScrapingJob.tenant_id == user.tenant_id,
        )
    )
    job = job_result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    result = await db.execute(
        select(ScrapedProfile)
        .where(ScrapedProfile.job_id == job_id, ScrapedProfile.scrape_status == "success")
    )
    profiles = result.scalars().all()

    output = io.StringIO()

    # Author info header
    author = await _get_author(db, job_id)
    if author:
        output.write(f"# Page: {author.name or ''} ({author.platform_object_id})")
        if author.category:
            output.write(f" | Category: {author.category}")
        if author.location:
            output.write(f" | Location: {author.location}")
        output.write("\n")

    writer = csv.DictWriter(output, fieldnames=FIELDNAMES)
    writer.writeheader()

    for p in profiles:
        writer.writerow({
            "ID": p.platform_user_id,
            "Name": p.name or "NA",
            "First Name": p.first_name or "NA",
            "Last Name": p.last_name or "NA",
            "Gender": p.gender or "NA",
            "Birthday": p.birthday or "NA",
            "Phone": p.phone or "NA",
            "Relationship": p.relationship_status or "NA",
            "Education": p.education or "NA",
            "Work": p.work or "NA",
            "Position": p.position or "NA",
            "Hometown": p.hometown or "NA",
            "Location": p.location or "NA",
            "Website": p.website or "NA",
            "Languages": p.languages or "NA",
            "UsernameLink": p.username_link or "NA",
            "Username": p.username or "NA",
            "About": p.about or "NA",
            "Picture URL": p.picture_url or "NA",
            "Updated Time": p.scraped_at.strftime("%Y-%m-%d %H:%M:%S") if p.scraped_at else "NA",
        })

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=socybase_fb_ads_{job_id}.csv"},
    )


# Post discovery export fields
POST_FIELDNAMES = [
    "Post ID", "Message", "Author", "Author ID", "Created",
    "Comments", "Reactions", "Shares", "Type", "URL",
]


@router.get("/{job_id}/xlsx")
async def export_xlsx(
    job_id: UUID,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export job data as XLSX. Supports both profile scraping and post discovery jobs."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    job_result = await db.execute(
        select(ScrapingJob).where(
            ScrapingJob.id == job_id,
            ScrapingJob.tenant_id == user.tenant_id,
        )
    )
    job = job_result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    if job.job_type == "post_discovery":
        ws.title = "Discovered Posts"
        headers = POST_FIELDNAMES

        result = await db.execute(
            select(ScrapedPost).where(ScrapedPost.job_id == job_id)
            .order_by(ScrapedPost.created_time.desc().nulls_last())
        )
        posts = result.scalars().all()

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, p in enumerate(posts, 2):
            ws.cell(row=row_idx, column=1, value=p.post_id)
            ws.cell(row=row_idx, column=2, value=(p.message or "")[:500])
            ws.cell(row=row_idx, column=3, value=p.from_name or "")
            ws.cell(row=row_idx, column=4, value=p.from_id or "")
            ws.cell(row=row_idx, column=5, value=p.created_time.strftime("%Y-%m-%d %H:%M:%S") if p.created_time else "")
            ws.cell(row=row_idx, column=6, value=p.comment_count)
            ws.cell(row=row_idx, column=7, value=p.reaction_count)
            ws.cell(row=row_idx, column=8, value=p.share_count)
            ws.cell(row=row_idx, column=9, value=p.attachment_type or "status")
            ws.cell(row=row_idx, column=10, value=p.post_url or "")
    else:
        ws.title = "Scraped Profiles"
        headers = FIELDNAMES

        result = await db.execute(
            select(ScrapedProfile)
            .where(ScrapedProfile.job_id == job_id, ScrapedProfile.scrape_status == "success")
        )
        profiles = result.scalars().all()

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, p in enumerate(profiles, 2):
            ws.cell(row=row_idx, column=1, value=p.platform_user_id)
            ws.cell(row=row_idx, column=2, value=p.name or "NA")
            ws.cell(row=row_idx, column=3, value=p.first_name or "NA")
            ws.cell(row=row_idx, column=4, value=p.last_name or "NA")
            ws.cell(row=row_idx, column=5, value=p.gender or "NA")
            ws.cell(row=row_idx, column=6, value=p.birthday or "NA")
            ws.cell(row=row_idx, column=7, value=p.phone or "NA")
            ws.cell(row=row_idx, column=8, value=p.relationship_status or "NA")
            ws.cell(row=row_idx, column=9, value=p.education or "NA")
            ws.cell(row=row_idx, column=10, value=p.work or "NA")
            ws.cell(row=row_idx, column=11, value=p.position or "NA")
            ws.cell(row=row_idx, column=12, value=p.hometown or "NA")
            ws.cell(row=row_idx, column=13, value=p.location or "NA")
            ws.cell(row=row_idx, column=14, value=p.website or "NA")
            ws.cell(row=row_idx, column=15, value=p.languages or "NA")
            ws.cell(row=row_idx, column=16, value=p.username_link or "NA")
            ws.cell(row=row_idx, column=17, value=p.username or "NA")
            ws.cell(row=row_idx, column=18, value=p.about or "NA")
            ws.cell(row=row_idx, column=19, value=p.picture_url or "NA")
            ws.cell(row=row_idx, column=20, value=p.scraped_at.strftime("%Y-%m-%d %H:%M:%S") if p.scraped_at else "NA")

    # Author Info sheet (if available)
    author = await _get_author(db, job_id)
    if author:
        ws_author = wb.create_sheet("Author Info")
        author_rows = [
            ("Field", "Value"),
            ("Page ID", author.platform_object_id),
            ("Name", author.name or ""),
            ("Category", author.category or ""),
            ("About", author.about or ""),
            ("Description", author.description or ""),
            ("Location", author.location or ""),
            ("Phone", author.phone or ""),
            ("Website", author.website or ""),
            ("Picture URL", author.picture_url or ""),
            ("Cover URL", author.cover_url or ""),
            ("Fetched At", author.fetched_at.strftime("%Y-%m-%d %H:%M:%S") if author.fetched_at else ""),
        ]
        for row_idx, (field, value) in enumerate(author_rows, 1):
            cell_f = ws_author.cell(row=row_idx, column=1, value=field)
            ws_author.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                cell_f.font = header_font
                cell_f.fill = header_fill
                ws_author.cell(row=1, column=2).font = header_font
                ws_author.cell(row=1, column=2).fill = header_fill
            else:
                cell_f.font = Font(bold=True)
        ws_author.column_dimensions["A"].width = 15
        ws_author.column_dimensions["B"].width = 60

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_length + 3

    xlsx_output = io.BytesIO()
    wb.save(xlsx_output)
    xlsx_output.seek(0)

    return StreamingResponse(
        iter([xlsx_output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=socybase_export_{job_id}.xlsx"},
    )


# ── Batch Export (ZIP) ─────────────────────────────────────────────


class BatchExportRequest(BaseModel):
    job_ids: list[str] = Field(..., min_length=1, max_length=50)
    format: str = Field(default="csv", pattern="^(csv|xlsx)$")


def _generate_csv_bytes(profiles, author) -> bytes:
    """Generate CSV content for a list of profiles."""
    output = io.StringIO()
    if author:
        output.write(f"# Page: {author.name or ''} ({author.platform_object_id})")
        if author.category:
            output.write(f" | Category: {author.category}")
        if author.location:
            output.write(f" | Location: {author.location}")
        output.write("\n")

    writer = csv.DictWriter(output, fieldnames=FIELDNAMES)
    writer.writeheader()
    for p in profiles:
        writer.writerow({
            "ID": p.platform_user_id,
            "Name": p.name or "NA",
            "First Name": p.first_name or "NA",
            "Last Name": p.last_name or "NA",
            "Gender": p.gender or "NA",
            "Birthday": p.birthday or "NA",
            "Phone": p.phone or "NA",
            "Relationship": p.relationship_status or "NA",
            "Education": p.education or "NA",
            "Work": p.work or "NA",
            "Position": p.position or "NA",
            "Hometown": p.hometown or "NA",
            "Location": p.location or "NA",
            "Website": p.website or "NA",
            "Languages": p.languages or "NA",
            "UsernameLink": p.username_link or "NA",
            "Username": p.username or "NA",
            "About": p.about or "NA",
            "Picture URL": p.picture_url or "NA",
            "Updated Time": p.scraped_at.strftime("%Y-%m-%d %H:%M:%S") if p.scraped_at else "NA",
        })
    return output.getvalue().encode("utf-8")


def _generate_post_csv_bytes(posts, author) -> bytes:
    """Generate CSV content for post discovery results."""
    output = io.StringIO()
    if author:
        output.write(f"# Page: {author.name or ''} ({author.platform_object_id})\n")

    writer = csv.DictWriter(output, fieldnames=POST_FIELDNAMES)
    writer.writeheader()
    for p in posts:
        writer.writerow({
            "Post ID": p.post_id,
            "Message": (p.message or "")[:500],
            "Author": p.from_name or "",
            "Author ID": p.from_id or "",
            "Created": p.created_time.strftime("%Y-%m-%d %H:%M:%S") if p.created_time else "",
            "Comments": p.comment_count,
            "Reactions": p.reaction_count,
            "Shares": p.share_count,
            "Type": p.attachment_type or "status",
            "URL": p.post_url or "",
        })
    return output.getvalue().encode("utf-8")


@router.post("/batch")
async def batch_export(
    data: BatchExportRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export multiple jobs as a ZIP file containing individual CSV/XLSX files."""
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for jid_str in data.job_ids:
            try:
                jid = UUID(jid_str)
            except ValueError:
                continue

            job_result = await db.execute(
                select(ScrapingJob).where(
                    ScrapingJob.id == jid,
                    ScrapingJob.tenant_id == user.tenant_id,
                )
            )
            job = job_result.scalar_one_or_none()
            if not job:
                continue

            author = await _get_author(db, jid)
            # Determine a short label for the filename
            label = (job.input_value or str(jid))[:40].replace("/", "_").replace("\\", "_").replace(":", "")

            if job.job_type == "post_discovery":
                result = await db.execute(
                    select(ScrapedPost).where(ScrapedPost.job_id == jid)
                    .order_by(ScrapedPost.created_time.desc().nulls_last())
                )
                posts = result.scalars().all()
                csv_bytes = _generate_post_csv_bytes(posts, author)
                zf.writestr(f"{label}_{str(jid)[:8]}.csv", csv_bytes)
            else:
                result = await db.execute(
                    select(ScrapedProfile)
                    .where(ScrapedProfile.job_id == jid, ScrapedProfile.scrape_status == "success")
                )
                profiles = result.scalars().all()
                csv_bytes = _generate_csv_bytes(profiles, author)
                zf.writestr(f"{label}_{str(jid)[:8]}.csv", csv_bytes)

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=socybase_batch_export.zip"},
    )
